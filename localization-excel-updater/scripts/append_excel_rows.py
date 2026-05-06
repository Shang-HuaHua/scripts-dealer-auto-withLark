#!/usr/bin/env python3
import argparse
import json
from pathlib import Path

from openpyxl import load_workbook


SIDE_DIRS = {
    "app": Path("app端"),
    "glasses": Path("眼镜端/applocalizationtool4anndroid"),
}

SIDE_ALIASES = {
    "app": "app",
    "app端": "app",
    "glasses": "glasses",
    "眼镜端": "glasses",
}


def parse_args():
    parser = argparse.ArgumentParser(
        description="Append localization rows to an Excel workbook after validating key uniqueness."
    )
    parser.add_argument("--workspace", required=True, help="Workspace root containing app端 and 眼镜端")
    parser.add_argument("--side", required=True, help="Target side: app, app端, glasses, or 眼镜端")
    parser.add_argument("--file", required=True, help="Excel module name with or without .xlsx")
    parser.add_argument(
        "--input",
        required=True,
        help="Path to a JSON file containing a list of row objects to append",
    )
    return parser.parse_args()


def normalize_module_name(file_name: str) -> str:
    return file_name[:-5] if file_name.endswith(".xlsx") else file_name


def resolve_excel_path(workspace: Path, side: str, file_name: str) -> Path:
    module = normalize_module_name(file_name)
    return workspace / SIDE_DIRS[side] / "excel_files" / f"{module}.xlsx"


def normalize_side(side: str) -> str:
    normalized = SIDE_ALIASES.get(side.strip())
    if normalized is None:
        supported = ", ".join(sorted(SIDE_ALIASES))
        raise ValueError(f"Unsupported side: {side}. Expected one of: {supported}")
    return normalized


def load_rows(input_path: Path):
    rows = json.loads(input_path.read_text(encoding="utf-8"))
    if not isinstance(rows, list) or not rows:
        raise ValueError("Input JSON must be a non-empty list")
    for index, row in enumerate(rows, start=1):
        if not isinstance(row, dict):
            raise ValueError(f"Row {index} must be an object")
        key = row.get("key")
        if not isinstance(key, str) or not key.strip():
            raise ValueError(f"Row {index} is missing a non-empty key")
        row["key"] = key.strip()
    return rows


def read_headers(worksheet):
    headers = []
    for cell in worksheet[1]:
        if cell.value is None:
            break
        headers.append(str(cell.value).strip())
    if not headers or headers[0] != "key":
        raise ValueError("Worksheet must start with a key column")
    return headers


def read_existing_keys(worksheet):
    keys = set()
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        value = row[0] if row else None
        if value is None:
            continue
        keys.add(str(value).strip())
    return keys


def find_last_nonempty_key_row(worksheet):
    last_row = 1
    for row_index, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        value = row[0] if row else None
        if value is None:
            continue
        if str(value).strip():
            last_row = row_index
    return last_row


def append_rows(excel_path: Path, rows):
    workbook = load_workbook(excel_path)
    worksheet = workbook[workbook.sheetnames[0]]

    headers = read_headers(worksheet)
    existing_keys = read_existing_keys(worksheet)
    last_nonempty_key_row = find_last_nonempty_key_row(worksheet)
    pending_keys = set()

    for index, row in enumerate(rows, start=1):
        key = row["key"]
        if key in existing_keys:
            raise ValueError(f"Row {index} key already exists in workbook: {key}")
        if key in pending_keys:
            raise ValueError(f"Row {index} key duplicates another pending row: {key}")
        pending_keys.add(key)

    start_row = last_nonempty_key_row + 1
    for offset, row in enumerate(rows):
        target_row = start_row + offset
        for column_index, header in enumerate(headers, start=1):
            worksheet.cell(row=target_row, column=column_index, value=row.get(header, ""))

    workbook.save(excel_path)
    return headers, start_row, start_row + len(rows) - 1


def main():
    try:
        args = parse_args()
        workspace = Path(args.workspace).expanduser().resolve()
        input_path = Path(args.input).expanduser().resolve()
        side = normalize_side(args.side)
        excel_path = resolve_excel_path(workspace, side, args.file)

        if not excel_path.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")
        if not input_path.exists():
            raise FileNotFoundError(f"Input JSON not found: {input_path}")

        rows = load_rows(input_path)
        headers, start_row, end_row = append_rows(excel_path, rows)

        print(json.dumps(
            {
                "excel_path": str(excel_path),
                "headers": headers,
                "row_count": len(rows),
                "start_row": start_row,
                "end_row": end_row,
                "keys": [row["key"] for row in rows],
            },
            ensure_ascii=False,
            indent=2,
        ))
    except Exception as exc:
        raise SystemExit(f"ERROR: {exc}")


if __name__ == "__main__":
    main()
