#!/usr/bin/env python3
import argparse
import json
from pathlib import Path

from openpyxl import load_workbook


SIDE_DIRS = {
    "app": Path("app端"),
    "glasses": Path("眼镜端/applocalizationtool4anndroid"),
}

SIDE_ALIASES = {
    "app": "app",
    "app端": "app",
    "glasses": "glasses",
    "眼镜端": "glasses",
}


def parse_args():
    parser = argparse.ArgumentParser(
        description="Update existing localization rows in an Excel workbook by key."
    )
    parser.add_argument("--workspace", required=True, help="Workspace root containing app端 and 眼镜端")
    parser.add_argument("--side", required=True, help="Target side: app, app端, glasses, or 眼镜端")
    parser.add_argument("--file", required=True, help="Excel module name with or without .xlsx")
    parser.add_argument(
        "--input",
        required=True,
        help="Path to a JSON file containing a list of row objects to update",
    )
    return parser.parse_args()


def normalize_module_name(file_name: str) -> str:
    return file_name[:-5] if file_name.endswith(".xlsx") else file_name


def normalize_side(side: str) -> str:
    normalized = SIDE_ALIASES.get(side.strip())
    if normalized is None:
        supported = ", ".join(sorted(SIDE_ALIASES))
        raise ValueError(f"Unsupported side: {side}. Expected one of: {supported}")
    return normalized


def resolve_excel_path(workspace: Path, side: str, file_name: str) -> Path:
    module = normalize_module_name(file_name)
    return workspace / SIDE_DIRS[side] / "excel_files" / f"{module}.xlsx"


def load_rows(input_path: Path):
    rows = json.loads(input_path.read_text(encoding="utf-8"))
    if not isinstance(rows, list) or not rows:
        raise ValueError("Input JSON must be a non-empty list")

    normalized_rows = []
    seen_keys = set()
    for index, row in enumerate(rows, start=1):
        if not isinstance(row, dict):
            raise ValueError(f"Row {index} must be an object")
        key = row.get("key")
        if not isinstance(key, str) or not key.strip():
            raise ValueError(f"Row {index} is missing a non-empty key")
        key = key.strip()
        if key in seen_keys:
            raise ValueError(f"Row {index} key duplicates another pending row: {key}")
        seen_keys.add(key)
        row["key"] = key
        normalized_rows.append(row)
    return normalized_rows


def read_headers(worksheet):
    headers = []
    for cell in worksheet[1]:
        if cell.value is None:
            break
        headers.append(str(cell.value).strip())
    if not headers or headers[0] != "key":
        raise ValueError("Worksheet must start with a key column")
    return headers


def index_existing_keys(worksheet):
    key_to_row = {}
    for row_index, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        value = row[0] if row else None
        if value is None:
            continue
        key = str(value).strip()
        if key:
            key_to_row[key] = row_index
    return key_to_row


def update_rows(excel_path: Path, rows):
    workbook = load_workbook(excel_path)
    worksheet = workbook[workbook.sheetnames[0]]

    headers = read_headers(worksheet)
    header_set = set(headers)
    key_to_row = index_existing_keys(worksheet)
    updated_rows = []

    for index, row in enumerate(rows, start=1):
        key = row["key"]
        target_row = key_to_row.get(key)
        if target_row is None:
            raise ValueError(f"Row {index} key does not exist in workbook: {key}")

        unknown_headers = sorted(set(row.keys()) - header_set)
        if unknown_headers:
            raise ValueError(f"Row {index} contains unknown headers: {unknown_headers}")

        for column_index, header in enumerate(headers, start=1):
            if header == "key":
                continue
            if header in row:
                worksheet.cell(row=target_row, column=column_index, value=row[header])

        updated_rows.append({"key": key, "row": target_row})

    workbook.save(excel_path)
    return headers, updated_rows


def main():
    try:
        args = parse_args()
        workspace = Path(args.workspace).expanduser().resolve()
        input_path = Path(args.input).expanduser().resolve()
        side = normalize_side(args.side)
        excel_path = resolve_excel_path(workspace, side, args.file)

        if not excel_path.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")
        if not input_path.exists():
            raise FileNotFoundError(f"Input JSON not found: {input_path}")

        rows = load_rows(input_path)
        headers, updated_rows = update_rows(excel_path, rows)

        print(json.dumps(
            {
                "excel_path": str(excel_path),
                "headers": headers,
                "row_count": len(rows),
                "updated_rows": updated_rows,
            },
            ensure_ascii=False,
            indent=2,
        ))
    except Exception as exc:
        raise SystemExit(f"ERROR: {exc}")


if __name__ == "__main__":
    main()
