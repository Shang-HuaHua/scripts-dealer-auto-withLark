#!/usr/bin/env python3
import argparse
import json
import os
import re
import urllib.parse
import urllib.request
from datetime import datetime
from pathlib import Path
from typing import Dict, Iterable, List, Optional

from openpyxl import Workbook, load_workbook


SIDE_ALIASES = {
    "app": "app",
    "app端": "app",
    "glasses": "glasses",
    "眼镜端": "glasses",
    "all": "all",
    "全部": "all",
    "两端": "all",
    "两端全部": "all",
}

SIDE_LABELS = {
    "app": "app端",
    "glasses": "眼镜端",
}

SIDE_DIRS = {
    "app": Path("app端"),
    "glasses": Path("眼镜端/applocalizationtool4anndroid"),
}


def parse_args():
    parser = argparse.ArgumentParser(
        description="Download or pull Feishu localization spreadsheets in bulk."
    )
    subparsers = parser.add_subparsers(dest="mode", required=True)

    download_parser = subparsers.add_parser(
        "download",
        help="Download Feishu spreadsheets into a local folder without touching the workspace Excel sources.",
    )
    add_common_scope_args(download_parser)
    download_parser.add_argument(
        "--workspace",
        help="Workspace root. Used only to derive the default download directory when --output-dir is omitted.",
    )
    download_parser.add_argument(
        "--output-dir",
        help="Directory to write downloaded .xlsx files into. Defaults to <workspace>/feishu_downloads/<timestamp>/",
    )

    pull_parser = subparsers.add_parser(
        "pull",
        help="Pull Feishu spreadsheets and overwrite local excel_files workbooks.",
    )
    add_common_scope_args(pull_parser)
    pull_parser.add_argument(
        "--workspace",
        required=True,
        help="Workspace root containing app端 and 眼镜端/applocalizationtool4anndroid",
    )

    create_parser = subparsers.add_parser(
        "create",
        help="Create a new local workbook and matching Feishu spreadsheet for one side.",
    )
    create_parser.add_argument(
        "--workspace",
        required=True,
        help="Workspace root containing app端 and 眼镜端/applocalizationtool4anndroid",
    )
    create_parser.add_argument(
        "--side",
        required=True,
        help="Target side: app, app端, glasses, or 眼镜端",
    )
    create_parser.add_argument(
        "--file",
        required=True,
        help="New workbook module name with or without .xlsx",
    )
    create_parser.add_argument(
        "--config",
        help="Optional path to feishu sync config JSON. Defaults to feishu_sync_config.json beside the skill.",
    )
    create_parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Plan the create operation without writing local or cloud workbooks.",
    )

    return parser.parse_args()


def add_common_scope_args(parser):
    parser.add_argument(
        "--scope",
        required=True,
        help="Target scope: app, app端, glasses, 眼镜端, all, 全部, or 两端全部",
    )
    parser.add_argument(
        "--file",
        action="append",
        dest="files",
        help="Optional workbook module name. Repeat to download or pull only selected modules.",
    )
    parser.add_argument(
        "--config",
        help="Optional path to feishu sync config JSON. Defaults to feishu_sync_config.json beside the skill.",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Plan the operation and print the matched files without writing any workbook.",
    )


def default_config_path() -> Path:
    return Path(__file__).resolve().parent.parent / "feishu_sync_config.json"


def load_config(config_path: Optional[str]):
    path = Path(config_path).expanduser().resolve() if config_path else default_config_path()
    config = {}
    if path.exists():
        config = json.loads(path.read_text(encoding="utf-8"))

    app_id = os.getenv("FEISHU_APP_ID") or config.get("app_id")
    app_secret = os.getenv("FEISHU_APP_SECRET") or config.get("app_secret")
    folders = dict(config.get("folders", {}))
    if os.getenv("FEISHU_FOLDER_APP"):
        folders["app"] = os.getenv("FEISHU_FOLDER_APP")
    if os.getenv("FEISHU_FOLDER_GLASSES"):
        folders["glasses"] = os.getenv("FEISHU_FOLDER_GLASSES")

    if not app_id or not app_secret:
        raise ValueError("Missing Feishu app_id/app_secret. Set them in config or env vars.")
    if not folders.get("app") or not folders.get("glasses"):
        raise ValueError("Missing Feishu folder tokens for app/glasses.")

    return {
        "app_id": app_id,
        "app_secret": app_secret,
        "folders": folders,
        "config_path": str(path),
    }


def normalize_scope(scope: str) -> str:
    normalized = SIDE_ALIASES.get(scope.strip())
    if normalized is None:
        supported = ", ".join(sorted(SIDE_ALIASES))
        raise ValueError(f"Unsupported scope: {scope}. Expected one of: {supported}")
    return normalized


def target_sides(scope: str) -> List[str]:
    return ["app", "glasses"] if scope == "all" else [scope]


def normalize_module_name(file_name: str) -> str:
    return file_name[:-5] if file_name.endswith(".xlsx") else file_name


MODULE_NAME_PATTERN = re.compile(r"^[A-Za-z]+(?:_[A-Za-z]+)*$")


def normalize_modules(files: Optional[Iterable[str]]) -> List[str]:
    if not files:
        return []
    return list(dict.fromkeys(normalize_module_name(file_name.strip()) for file_name in files if file_name.strip()))


def validate_scope_and_files(scope: str, files: List[str]):
    if scope == "all" and files:
        raise ValueError("When scope is all/全部, do not pass --file. Run each side separately for partial modules.")


def api_request(method: str, url: str, token: Optional[str] = None, json_body=None):
    headers = {}
    data = None
    if token:
        headers["Authorization"] = f"Bearer {token}"
    if json_body is not None:
        headers["Content-Type"] = "application/json"
        data = json.dumps(json_body).encode("utf-8")
    request = urllib.request.Request(url, data=data, headers=headers, method=method)
    with urllib.request.urlopen(request, timeout=60) as response:
        return json.loads(response.read().decode("utf-8"))


def get_tenant_access_token(config):
    data = api_request(
        "POST",
        "https://open.feishu.cn/open-apis/auth/v3/tenant_access_token/internal",
        json_body={"app_id": config["app_id"], "app_secret": config["app_secret"]},
    )
    if data.get("code") != 0:
        raise ValueError(f"Failed to get Feishu tenant access token: {data}")
    return data["tenant_access_token"]


def list_folder_files(token: str, folder_token: str):
    files = []
    page_token = ""
    while True:
        url = f"https://open.feishu.cn/open-apis/drive/v1/files?folder_token={folder_token}&page_size=200"
        if page_token:
            url += f"&page_token={urllib.parse.quote(page_token, safe='')}"
        data = api_request("GET", url, token=token)
        if data.get("code") != 0:
            raise ValueError(f"Failed to list Feishu folder: {data}")
        data_block = data.get("data", {})
        files.extend(data_block.get("files", []))
        if not data_block.get("has_more"):
            break
        page_token = data_block.get("next_page_token") or data_block.get("page_token") or ""
        if not page_token:
            break
    return files


def list_spreadsheets_for_side(token: str, folder_token: str):
    return [file for file in list_folder_files(token, folder_token) if file.get("type") == "sheet"]


def build_spreadsheet_index(files):
    exact = {}
    lowered = {}
    for file in files:
        name = str(file.get("name", "")).strip()
        if not name:
            continue
        exact[name] = file
        lowered[name.lower()] = file
    return exact, lowered


def resolve_spreadsheet(files, module_name: str):
    exact, lowered = build_spreadsheet_index(files)
    spreadsheet = exact.get(module_name)
    if spreadsheet:
        return spreadsheet
    spreadsheet = lowered.get(module_name.lower())
    if spreadsheet:
        return spreadsheet
    return None


def validate_module_name(module_name: str):
    if not module_name:
        raise ValueError("Workbook name cannot be empty.")
    if not MODULE_NAME_PATTERN.fullmatch(module_name):
        raise ValueError(
            "Workbook name must use English letters only, with underscores allowed only between words. "
            "Spaces, Chinese characters, digits, and other symbols are not allowed."
        )


def ensure_module_name_not_used(workspace_root: Path, side: str, module_name: str):
    existing_names = {name.lower(): path for name, path in local_workbook_map(workspace_root, side).items()}
    existing_path = existing_names.get(module_name.lower())
    if existing_path is not None:
        raise FileExistsError(
            f"Workbook name '{module_name}' conflicts with an existing {SIDE_LABELS[side]} Excel file: {existing_path}"
        )


def get_spreadsheet_metainfo(token: str, spreadsheet_token: str):
    data = api_request(
        "GET",
        f"https://open.feishu.cn/open-apis/sheets/v2/spreadsheets/{spreadsheet_token}/metainfo",
        token=token,
    )
    if data.get("code") != 0:
        raise ValueError(f"Failed to get spreadsheet metainfo: {data}")
    return data["data"]


def read_sheet_range(token: str, spreadsheet_token: str, range_expr: str):
    encoded = urllib.parse.quote(range_expr, safe="")
    data = api_request(
        "GET",
        f"https://open.feishu.cn/open-apis/sheets/v2/spreadsheets/{spreadsheet_token}/values/{encoded}",
        token=token,
    )
    if data.get("code") != 0:
        raise ValueError(f"Failed to read sheet range {range_expr}: {data}")
    return data["data"]["valueRange"].get("values", [])


def column_letter(index: int) -> str:
    letters = []
    while index > 0:
        index, remainder = divmod(index - 1, 26)
        letters.append(chr(65 + remainder))
    return "".join(reversed(letters))


def trim_matrix(values):
    if not values:
        return []

    normalized = []
    last_nonempty_row = 0
    last_nonempty_col = 0

    for row_index, row in enumerate(values, start=1):
        row_values = list(row) if isinstance(row, list) else [row]
        normalized.append(row_values)
        for col_index, cell in enumerate(row_values, start=1):
            if cell is None:
                continue
            if isinstance(cell, str):
                if cell.strip() == "":
                    continue
            last_nonempty_row = row_index
            if col_index > last_nonempty_col:
                last_nonempty_col = col_index

    if last_nonempty_row == 0 or last_nonempty_col == 0:
        return []

    trimmed = []
    for row in normalized[:last_nonempty_row]:
        row_values = row[:last_nonempty_col]
        if len(row_values) < last_nonempty_col:
            row_values = row_values + [""] * (last_nonempty_col - len(row_values))
        trimmed.append(["" if cell is None else cell for cell in row_values])
    return trimmed


def fetch_spreadsheet_values(token: str, spreadsheet):
    metainfo = get_spreadsheet_metainfo(token, spreadsheet["token"])
    first_sheet = metainfo["sheets"][0]
    last_col = column_letter(first_sheet["columnCount"])
    values = read_sheet_range(
        token,
        spreadsheet["token"],
        f"{first_sheet['sheetId']}!A1:{last_col}{first_sheet['rowCount']}",
    )
    trimmed_values = trim_matrix(values)
    if not trimmed_values or not trimmed_values[0] or str(trimmed_values[0][0]).strip() != "key":
        raise ValueError(
            f"Spreadsheet '{spreadsheet.get('name', spreadsheet['token'])}' does not start with a key header row."
        )
    return {
        "sheet_id": first_sheet["sheetId"],
        "sheet_title": first_sheet["title"],
        "row_count": len(trimmed_values),
        "column_count": max(len(row) for row in trimmed_values) if trimmed_values else 0,
        "values": trimmed_values,
    }


def sanitize_sheet_title(title: str) -> str:
    invalid = set('[]:*?/\\')
    cleaned = "".join("_" if char in invalid else char for char in title).strip()
    return (cleaned or "Sheet1")[:31]


def write_new_workbook(output_path: Path, sheet_title: str, values):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sanitize_sheet_title(sheet_title)
    for row_index, row in enumerate(values, start=1):
        for column_index, value in enumerate(row, start=1):
            worksheet.cell(row=row_index, column=column_index, value=value)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def read_local_headers_template(workspace_root: Path, side: str):
    excel_dir = local_excel_dir(workspace_root, side)
    candidates = sorted(excel_dir.glob("*.xlsx"), key=lambda item: item.stem.lower())
    for path in candidates:
        if path.stem.startswith(".~") or path.stem.startswith("~$"):
            continue
        workbook = load_workbook(path, data_only=True)
        worksheet = workbook[workbook.sheetnames[0]]
        headers = []
        for column_index in range(1, worksheet.max_column + 1):
            value = worksheet.cell(1, column_index).value
            if value is None:
                break
            headers.append(str(value).strip())
        if headers and headers[0] == "key":
            return headers
    raise ValueError(f"Could not infer workbook headers for side {side} from {excel_dir}")


def replace_local_workbook(local_path: Path, sheet_title: str, values):
    local_path.parent.mkdir(parents=True, exist_ok=True)
    temp_path = local_path.with_name(f".{local_path.stem}.feishu_pull_tmp.xlsx")
    if temp_path.exists():
        temp_path.unlink()
    write_new_workbook(temp_path, sheet_title, values)
    temp_path.replace(local_path)


def default_download_output_dir(workspace_root: Path) -> Path:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return workspace_root / "feishu_downloads" / timestamp


def resolve_workspace_root(workspace: Optional[str]) -> Path:
    return Path(workspace).expanduser().resolve() if workspace else Path.cwd().resolve()


def collect_download_targets(files, requested_modules: List[str]):
    if not requested_modules:
        return sorted(files, key=lambda item: str(item.get("name", "")).lower())

    resolved = []
    missing = []
    for module in requested_modules:
        spreadsheet = resolve_spreadsheet(files, module)
        if spreadsheet is None:
            missing.append(module)
            continue
        resolved.append(spreadsheet)
    if missing:
        raise FileNotFoundError(f"Requested Feishu spreadsheets not found: {sorted(missing)}")
    return resolved


def run_download(args):
    scope = normalize_scope(args.scope)
    requested_modules = normalize_modules(args.files)
    validate_scope_and_files(scope, requested_modules)

    workspace_root = resolve_workspace_root(args.workspace)
    output_root = Path(args.output_dir).expanduser().resolve() if args.output_dir else default_download_output_dir(workspace_root)
    config = load_config(args.config)
    token = get_tenant_access_token(config)

    results = []
    total_downloaded = 0

    for side in target_sides(scope):
        spreadsheets = list_spreadsheets_for_side(token, config["folders"][side])
        targets = collect_download_targets(spreadsheets, requested_modules)
        side_output_root = output_root / SIDE_LABELS[side]
        downloaded = []

        for spreadsheet in targets:
            sheet_data = fetch_spreadsheet_values(token, spreadsheet)
            output_path = side_output_root / f"{spreadsheet['name']}.xlsx"
            if not args.dry_run:
                write_new_workbook(output_path, sheet_data["sheet_title"], sheet_data["values"])
            downloaded.append(
                {
                    "module": spreadsheet["name"],
                    "output_path": str(output_path),
                    "spreadsheet_token": spreadsheet["token"],
                    "spreadsheet_url": spreadsheet.get("url"),
                    "sheet_id": sheet_data["sheet_id"],
                    "sheet_title": sheet_data["sheet_title"],
                    "row_count": sheet_data["row_count"],
                    "column_count": sheet_data["column_count"],
                }
            )

        total_downloaded += len(downloaded)
        results.append(
            {
                "side": side,
                "side_label": SIDE_LABELS[side],
                "folder_token": config["folders"][side],
                "download_root": str(side_output_root),
                "requested_modules": requested_modules,
                "downloaded": downloaded,
            }
        )

    if total_downloaded == 0:
        raise ValueError("No spreadsheets were downloaded.")

    print(
        json.dumps(
            {
                "mode": "download",
                "scope": scope,
                "output_root": str(output_root),
                "dry_run": args.dry_run,
                "total_downloaded": total_downloaded,
                "results": results,
            },
            ensure_ascii=False,
            indent=2,
        )
    )


def local_excel_dir(workspace_root: Path, side: str) -> Path:
    return workspace_root / SIDE_DIRS[side] / "excel_files"


def local_workbook_map(workspace_root: Path, side: str) -> Dict[str, Path]:
    return {
        path.stem: path
        for path in sorted(local_excel_dir(workspace_root, side).glob("*.xlsx"))
        if not path.stem.startswith(".~") and not path.stem.startswith("~$")
    }


def create_feishu_spreadsheet(token: str, folder_token: str, module_name: str):
    data = api_request(
        "POST",
        "https://open.feishu.cn/open-apis/sheets/v3/spreadsheets",
        token=token,
        json_body={"title": module_name, "folder_token": folder_token},
    )
    if data.get("code") != 0:
        raise ValueError(f"Failed to create Feishu spreadsheet: {data}")
    return data["data"]["spreadsheet"]


def write_feishu_headers(token: str, spreadsheet_token: str, headers: List[str]):
    metainfo = get_spreadsheet_metainfo(token, spreadsheet_token)
    first_sheet = metainfo["sheets"][0]
    range_expr = f"{first_sheet['sheetId']}!A1:{column_letter(len(headers))}1"
    data = api_request(
        "PUT",
        f"https://open.feishu.cn/open-apis/sheets/v2/spreadsheets/{spreadsheet_token}/values",
        token=token,
        json_body={"valueRange": {"range": range_expr, "values": [headers]}},
    )
    if data.get("code") != 0:
        raise ValueError(f"Failed to write Feishu headers: {data}")
    return {
        "sheet_id": first_sheet["sheetId"],
        "sheet_title": first_sheet["title"],
        "updated_range": data["data"].get("updatedRange"),
    }


def run_pull(args):
    scope = normalize_scope(args.scope)
    requested_modules = normalize_modules(args.files)
    validate_scope_and_files(scope, requested_modules)

    workspace_root = Path(args.workspace).expanduser().resolve()
    config = load_config(args.config)
    token = get_tenant_access_token(config)

    results = []
    total_synced = 0

    for side in target_sides(scope):
        spreadsheets = list_spreadsheets_for_side(token, config["folders"][side])
        local_map = local_workbook_map(workspace_root, side)

        if requested_modules:
            missing_local = [module for module in requested_modules if module not in local_map]
            missing_cloud = [module for module in requested_modules if resolve_spreadsheet(spreadsheets, module) is None]
            if missing_local:
                raise FileNotFoundError(
                    f"Requested local workbook(s) not found under {local_excel_dir(workspace_root, side)}: {sorted(missing_local)}"
                )
            if missing_cloud:
                raise FileNotFoundError(f"Requested Feishu spreadsheet(s) not found: {sorted(missing_cloud)}")
            target_modules = requested_modules
        else:
            target_modules = sorted(local_map)

        synced = []
        missing_in_cloud = []
        for module in target_modules:
            local_path = local_map.get(module)
            spreadsheet = resolve_spreadsheet(spreadsheets, module)
            if local_path is None:
                continue
            if spreadsheet is None:
                missing_in_cloud.append(module)
                continue

            sheet_data = fetch_spreadsheet_values(token, spreadsheet)
            if not args.dry_run:
                replace_local_workbook(local_path, sheet_data["sheet_title"], sheet_data["values"])
            synced.append(
                {
                    "module": module,
                    "local_path": str(local_path),
                    "spreadsheet_token": spreadsheet["token"],
                    "spreadsheet_url": spreadsheet.get("url"),
                    "sheet_id": sheet_data["sheet_id"],
                    "sheet_title": sheet_data["sheet_title"],
                    "row_count": sheet_data["row_count"],
                    "column_count": sheet_data["column_count"],
                }
            )

        cloud_names = {str(file.get("name", "")).strip() for file in spreadsheets if str(file.get("name", "")).strip()}
        local_names = set(local_map)
        extra_in_cloud = sorted(name for name in cloud_names if name not in local_names)

        total_synced += len(synced)
        results.append(
            {
                "side": side,
                "side_label": SIDE_LABELS[side],
                "folder_token": config["folders"][side],
                "local_excel_dir": str(local_excel_dir(workspace_root, side)),
                "requested_modules": requested_modules,
                "synced": synced,
                "missing_in_cloud": sorted(missing_in_cloud),
                "extra_in_cloud": extra_in_cloud,
            }
        )

    if total_synced == 0:
        raise ValueError("No local workbooks were overwritten from Feishu.")

    print(
        json.dumps(
            {
                "mode": "pull",
                "scope": scope,
                "workspace": str(workspace_root),
                "dry_run": args.dry_run,
                "total_synced": total_synced,
                "results": results,
            },
            ensure_ascii=False,
            indent=2,
        )
    )


def run_create(args):
    workspace_root = Path(args.workspace).expanduser().resolve()
    side = normalize_scope(args.side)
    if side == "all":
        raise ValueError("Create mode supports only one side at a time.")

    module_name = normalize_module_name(args.file)
    validate_module_name(module_name)
    ensure_module_name_not_used(workspace_root, side, module_name)

    config = load_config(args.config)
    token = get_tenant_access_token(config)

    headers = read_local_headers_template(workspace_root, side)
    target_folder = local_excel_dir(workspace_root, side).resolve()
    local_path = target_folder / f"{module_name}.xlsx"

    spreadsheets = list_spreadsheets_for_side(token, config["folders"][side])
    existing_cloud = resolve_spreadsheet(spreadsheets, module_name)
    if existing_cloud is not None:
        raise FileExistsError(
            f"Feishu spreadsheet already exists for module '{module_name}': {existing_cloud.get('url')}"
        )

    cloud_result = None
    if not args.dry_run:
        write_new_workbook(local_path, "Sheet1", [headers])
        created = create_feishu_spreadsheet(token, config["folders"][side], module_name)
        header_write = write_feishu_headers(token, created["spreadsheet_token"], headers)
        cloud_result = {
            "module": module_name,
            "spreadsheet_token": created["spreadsheet_token"],
            "spreadsheet_url": created["url"],
            "sheet_id": header_write["sheet_id"],
            "sheet_title": header_write["sheet_title"],
            "updated_range": header_write["updated_range"],
        }

    print(
        json.dumps(
            {
                "mode": "create",
                "side": side,
                "side_label": SIDE_LABELS[side],
                "module": module_name,
                "local_path": str(local_path),
                "target_folder": str(target_folder),
                "headers": headers,
                "folder_token": config["folders"][side],
                "dry_run": args.dry_run,
                "cloud": cloud_result,
            },
            ensure_ascii=False,
            indent=2,
        )
    )


def main():
    try:
        args = parse_args()
        if args.mode == "download":
            run_download(args)
        elif args.mode == "pull":
            run_pull(args)
        elif args.mode == "create":
            run_create(args)
        else:
            raise ValueError(f"Unsupported mode: {args.mode}")
    except Exception as exc:
        raise SystemExit(f"ERROR: {exc}")


if __name__ == "__main__":
    main()
